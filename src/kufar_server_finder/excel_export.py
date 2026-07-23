from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from kufar_finder_core import load_items

MISSING_VALUE = "—"

HEADERS = (
    "Тип устройства",
    "Цена, BYN",
    "Тип ОЗУ",
    "Объём ОЗУ, ГБ",
    "Сокет процессора",
    "Модель процессора",
    "CPU Benchmark",
    "Оценочная мощность системы, Вт",
    "Ссылка",
)

PRODUCT_TYPE_NAMES = {
    "desktop_pc": "Настольный ПК",
    "laptop": "Ноутбук",
    "mini_pc": "Мини-ПК",
    "thin_client": "Тонкий клиент",
    "server": "Сервер",
    "workstation": "Рабочая станция",
    "all_in_one": "Моноблок",
    "motherboard_bundle": "Комплект материнской платы",
    "other": "Другое",
}

CONFIDENCE_FILLS = {
    "low": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "medium": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "high": PatternFill(fill_type="solid", fgColor="C6EFCE"),
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")

FIELD_COLUMNS = {
    "product_type": 1,
    "ram_type": 3,
    "ram_gb": 4,
    "cpu_socket": 5,
    "cpu_model": 6,
    "cpu_mark": 7,
    "estimated_system_power_w": 8,
}

COLUMN_WIDTHS = {
    1: 24,
    2: 13,
    3: 14,
    4: 18,
    5: 20,
    6: 30,
    7: 16,
    8: 34,
    9: 55,
}


def export_ads_json_to_excel(
    json_path: str | Path,
    excel_path: str | Path,
) -> None:
    """Преобразует JSON-массив объявлений в оформленный Excel-документ."""
    ads = load_items(json_path)
    export_ads_to_excel(ads, excel_path)


def export_ads_to_excel(
    ads: list[dict[str, Any]],
    excel_path: str | Path,
) -> None:
    destination = Path(excel_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Объявления"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(len(ads) + 1, 1)}"

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, ad in enumerate(ads, start=2):
        values = (
            _display_product_type(ad.get("product_type")),
            _value_or_dash(ad.get("price")),
            _value_or_dash(ad.get("ram_type")),
            _value_or_dash(ad.get("ram_gb")),
            _value_or_dash(ad.get("cpu_socket")),
            _value_or_dash(ad.get("cpu_model")),
            _value_or_dash(ad.get("cpu_mark")),
            _value_or_dash(ad.get("estimated_system_power_w")),
            _value_or_dash(ad.get("link")),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.cell(row=row_index, column=2).number_format = "0.00"
        sheet.cell(row=row_index, column=4).number_format = "0"
        sheet.cell(row=row_index, column=7).number_format = "0.##"
        sheet.cell(row=row_index, column=8).number_format = "0"

        _format_link(sheet.cell(row=row_index, column=9), ad.get("link"))
        _apply_confidence_colors(sheet, row_index, ad)

    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[1].height = 28

    workbook.save(destination)


def _apply_confidence_colors(sheet: Any, row: int, ad: dict[str, Any]) -> None:
    for field, column in FIELD_COLUMNS.items():
        value = ad.get(field)
        if value in (None, ""):
            continue

        confidence = ad.get(f"{field}_confidence")
        if confidence not in CONFIDENCE_FILLS:
            source = ad.get(f"{field}_source")
            if source == "text_exact" or (
                field == "cpu_mark" and ad.get("cpu_benchmark_source") == "dataset"
            ):
                confidence = "high"

        fill = CONFIDENCE_FILLS.get(confidence)
        if fill is not None:
            sheet.cell(row=row, column=column).fill = fill


def _format_link(cell: Cell, link: Any) -> None:
    if not isinstance(link, str) or not link.startswith(("http://", "https://")):
        return
    cell.hyperlink = link
    cell.font = LINK_FONT


def _display_product_type(value: Any) -> Any:
    if value in (None, ""):
        return MISSING_VALUE
    return PRODUCT_TYPE_NAMES.get(str(value), value)


def _value_or_dash(value: Any) -> Any:
    return MISSING_VALUE if value in (None, "") else value
